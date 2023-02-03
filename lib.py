<<<<<<< Updated upstream
import librosa
import matplotlib.pyplot as plt
import librosa.display
import numpy as np

SEGMENT_NUM = 8

def cal_collerate(onsetS):
    # 8 seconds = 344   2 seconds = 86  10 seconds = 430  7.5 seconds = 322
    result = []
    init_index = 0
    while (init_index + 430) < len(onsetS):

        origin = onsetS[init_index : init_index + 344]
        similarity_list = []

        for i in range(86):
            target = onsetS[init_index + i : init_index + 344 + i]
            similarity = np.correlate(origin, target)
            similarity_list.append(similarity[0])
        
        result.append(similarity_list)
        init_index = init_index + 322

    return result

def chroma_segment(audio,sample_rate):
    X = librosa.feature.chroma_cqt(y=audio, sr=sample_rate)
    print('x:',len(X),'y:',len(X[0]))
    bounds = librosa.segment.agglomerative(X, SEGMENT_NUM)
    bound_times = librosa.frames_to_time(bounds, sr=sample_rate)
    return bound_times

def mfcc_segment(audio,sample_rate):
    X = librosa.feature.mfcc(y=audio, sr=sample_rate)
    print('x:',len(X),'y:',len(X[0]))
    bounds = librosa.segment.agglomerative(X, SEGMENT_NUM)
    bound_times = librosa.frames_to_time(bounds, sr=sample_rate)
    return bound_times


def rhythm_segment(audio,sample_rate):
    o_env = librosa.onset.onset_strength(y=audio, sr=sample_rate)

    print(len(o_env))

    collerate_list = cal_collerate(o_env)
    
    collerate_list_trans = list(map(list,zip(*collerate_list)))

    #print('x:',len(collerate_list),'y:',len(collerate_list[0]))

    bounds = librosa.segment.agglomerative(collerate_list_trans, SEGMENT_NUM)
    new_bounds = [i * 322 for i in bounds]
    bound_times = librosa.frames_to_time(new_bounds, sr=sample_rate)
    return bound_times

=======
import librosa
import matplotlib.pyplot as plt
import librosa.display
import numpy as np
import datetime
import os
import xlrd
import xlsxwriter
from openpyxl import load_workbook

#SEGMENT_NUM = 6


# create data sheet
def create_table(path, labels=None, sheet=None):
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet()
    if sheet is not None:
        worksheet.set_vba_name(sheet)
    if labels is not None:
        worksheet.write_row('A1', labels)
    return workbook, worksheet


def write_data(data, worksheet, start_row=0, start_column=0):
    for i in range(len(data)):
        worksheet.write_row(start_row + i, start_column, data[i])


def get_all_data(filepath, sheet='Sheet1'):
    wb = load_workbook(filepath)
    ws = wb[sheet]
    result = []

    for row in ws.rows:
        row_cells = []
        for cell in row:
            row_cells.append(cell.value)
        result.append(row_cells)

    return result


def find_all_song_file(database):
    for root, dirs, files in os.walk(database):
        for file in files:
            if file.endswith('.mp3'):
                fullname_list = os.path.join(file)
                yield fullname_list


def excel_read(excel, sheet_num):
    result = {}

    workbook = xlrd.open_workbook(excel)
    table = workbook.sheet_by_index(sheet_num)

    for i in range(table.nrows):

        song_segment = []
        song_name = ''

        for j in range(table.ncols):

            # song name
            if j == 0:
                song_name = table.cell(i, j).value
                continue

            if table.cell(i, j).value == '':
                break

            data = table.cell(i, j).value

            if j % 2 == 0:
                data = float(table.cell(i, j).value)
                data = round(data * 24 * 60)

            song_segment.append(data)

        result[song_name] = song_segment

    return result


def get_excel_data_times(excel_data):
    result = []
    for i in range(len(excel_data)):
        if i % 2 == 1:
            result.append(excel_data[i])

    return result


def get_result_sequence(data, duration):
    result = np.zeros(int(np.floor(duration)) + 1)
    for i in data:
        result[int(np.floor(i))] = 1
    return result


def cal_collerate(onsetS):
    # 8 seconds = 344   2 seconds = 86  10 seconds = 430  7.5 seconds = 322
    result = []
    init_index = 0
    while (init_index + 430) < len(onsetS):

        origin = onsetS[init_index: init_index + 344]
        similarity_list = []

        for i in range(86):
            target = onsetS[init_index + i: init_index + 344 + i]
            similarity = np.correlate(origin, target)
            similarity_list.append(similarity[0])

        result.append(similarity_list)
        init_index = init_index + 322

    return result

def cal_collerate_2(onsetS):
    # 8 seconds = 344   2 seconds = 86  10 seconds = 430  7.5 seconds = 322  6seconds = 258  4second = 172
    result = []
    init_index = 0
    while (init_index + 258) < len(onsetS):

        origin = onsetS[init_index: init_index + 172]
        similarity_list = []

        for i in range(86):
            target = onsetS[init_index + i: init_index + 172 + i]
            similarity = np.correlate(origin, target)
            similarity_list.append(similarity[0])

        result.append(similarity_list)
        init_index = init_index + 150

    return result


def chroma_segment(audio, sample_rate, seg_num):
    X = librosa.feature.chroma_cqt(y=audio, sr=sample_rate)
    #X = librosa.feature.chroma_stft(y=audio, sr=sample_rate)
    #print('x:', len(X), 'y:', len(X[0]))
    bounds = librosa.segment.agglomerative(X, seg_num)
    bound_times = librosa.frames_to_time(bounds, sr=sample_rate)
    return bound_times


def mfcc_segment(audio, sample_rate, seg_num):
    X = librosa.feature.mfcc(y=audio, sr=sample_rate)
    #print('x:', len(X), 'y:', len(X[0]))
    bounds = librosa.segment.agglomerative(X, seg_num)
    bound_times = librosa.frames_to_time(bounds, sr=sample_rate)
    return bound_times


def rhythm_segment(audio, sample_rate, seg_num):
    o_env = librosa.onset.onset_strength(y=audio, sr=sample_rate)
    #collerate_list = cal_collerate(o_env)
    collerate_list = cal_collerate_2(o_env)
    collerate_list_trans = list(map(list, zip(*collerate_list)))

    # print('x:',len(collerate_list),'y:',len(collerate_list[0]))

    bounds = librosa.segment.agglomerative(collerate_list_trans, seg_num)
    #new_bounds = [i * 322 for i in bounds]
    new_bounds = [i * 150 for i in bounds]
    bound_times = librosa.frames_to_time(new_bounds, sr=sample_rate)
    return bound_times


def second_2_min(list_second):
    min_list = []
    for i in list_second:
        transform_i = datetime.timedelta(seconds=i)
        # tranform_i.microseconds = 0

        min_list.append(str(transform_i))
        # m, s = divmod(i, 60)
        # min_list.append((round(m),round(s)))
    return min_list

>>>>>>> Stashed changes
